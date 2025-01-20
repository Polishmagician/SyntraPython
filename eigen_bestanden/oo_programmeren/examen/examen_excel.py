from openpyxl import Workbook, load_workbook
from tabulate import tabulate

bestand = 'consultants_data.xlsx'


# Open Excel bestand en lees in:
def lees_excel(bestand):
    wb_sheet_start = load_workbook(bestand)
    sheet_start = wb_sheet_start.active
    headers = [cell.value for cell in sheet_start[1]]
    data = []
    for row in sheet_start.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    return data, headers


def toon_data(data, headers):
    print(tabulate(data,headers=headers))


def voeg_data_toe(data, headers):
    new_row = []
    new_id = data[-1][0] + 1
    new_row.append(new_id)
    for f in headers[1:]:
        keuze = input(f"Wat is {f}? ")
        new_row.append(keuze)
    new_row[5] = float(new_row[5])
    data.append(new_row)
    print("Consultant is toegevoegd!")
    return data


def verwijder_rij(data, headers):
    toon_data(data, headers)
    keuze = int(input("Welke ID wil je verwijderen?"))
    viable = check_id(data, keuze)
    if viable == 0:
        print(f"{keuze} not in list")
    else:
        data.pop(keuze - 1)
        print("Consultant is verwijderd!")
    return data


def pas_uurtarief_aan(data, headers):
    toon_data(data, headers)
    keuze = int(input("Van welke ID wil je het uurtarief aanpassen?"))
    viable = check_id(data, keuze)
    if viable == 0:
        print(f"{keuze} not in list")
    else:
        new_uurtarief = float(input(f"Wat is het nieuwe uurtarief van ID {keuze}? "))
        for row in data:
            if row[0] == keuze:
                row[5] = new_uurtarief
        print(f"Uurtarief van {keuze} is veranderd in {new_uurtarief}")
        return data


def sorteer_uurtarief(data, headers):
    t = sorted(data, key=lambda x: x[5], reverse=True)
    toon_data(t, headers)


def consultants_HR(data, headers):
    filter_data = []
    for f in data:
        if f[4] == "HR":
            filter_data.append(f)
    wegschrijven_excel(headers, filter_data)
    print("Excel aangemaakt")


def wegschrijven_excel(headers, source_data):
    # Create a new workbook for the combined data
    wb_combined = Workbook()
    combined_sheet = wb_combined.active
    filter_bestand = "consultants_HR.xlsx"
    combined_sheet.append(headers)
    for row in source_data:
        combined_sheet.append(list(row))
    wb_combined.save(filter_bestand)


def check_id(data, keuze):
    idlist = []
    for i, value in enumerate(data):
        idlist.append(data[i][0])
    if keuze in idlist:
        return 1
    else:
        return 0


data, headers = lees_excel(bestand)

functions_dict = {"1": toon_data, "2": voeg_data_toe, "3": verwijder_rij, "4": pas_uurtarief_aan, "5": sorteer_uurtarief, "6": consultants_HR}

while True:
    print("1: Toon data" + "\n"
                           "2: Voeg data toe" + "\n"
                                                "3: Verwijder rij" + "\n"
                                                                     "4: Pas uurtarief aan" + "\n"
                                                                                                           "5: Sorteer uurtarief" + "\n"
                                                                                                                                            "6: Schrijf consultantsHR weg naar excel" + "\n"
                                                                                                                                                                                                "STOP: Stop controller")
    keuze = input("Wat wil je doen?").upper()
    if keuze == "STOP":
        break
    elif keuze in functions_dict.keys():
        functions_dict[keuze](data, headers)
    else:
        print("Foute ingave!")

