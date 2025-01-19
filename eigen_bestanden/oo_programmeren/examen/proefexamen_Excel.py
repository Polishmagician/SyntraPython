from openpyxl import Workbook, load_workbook

bestand = 'teambuilding.xlsx'


#Open Excel bestand en lees in:
def lees_excel(bestand):
    wb_sheet_start =load_workbook(bestand)
    sheet_start = wb_sheet_start.active
    headers = [cell.value for cell in sheet_start[1]]
    data = []
    for row in sheet_start.iter_rows(min_row=2,values_only=True):
        data.append(list(row))
    return data,headers

def toon_data(data,headers):
    for f in data:
        print(*f)

def voeg_data_toe(data,headers):
    new_row = []
    new_id = data[-1][0]+1
    new_row.append(new_id)
    for f in headers[1:]:
        keuze = input(f"Wat is {f}? ")
        new_row.append(keuze)
    new_row[4] = int(new_row[4])
    data.append(new_row)
    return data

def verwijder_rij(data,headers):
    toon_data(data,headers)
    keuze = int(input("Welke ID wil je verwijderen?"))
    idlist = []
    for i,value in enumerate(data):
        idlist.append(data[i][0])
    if keuze in idlist:
        data.pop(keuze-1)
    else:
        print(f"{keuze} not in list")
    return data

def pas_locatie_aan(data,headers):
    toon_data(data,headers)
    keuze = int(input("Van welke ID wil je de locatie aanpassen?"))
    viable = check_id(data,keuze)
    if viable == 0:
        print(f"{keuze} not in list")
    else:
        new_location = input(f"Wat is de nieuwe locatie van ID {keuze}? ")
        for row in data:
            if row[0] == keuze:
                row[3] = new_location
        print(f"Activiteit van {keuze} is veranderd in {new_location}")
        return data

def pas_activiteit_aan(data,headers):
    toon_data(data,headers)
    keuze = int(input("Van welke ID wil je de activiteit aanpassen?"))
    viable = check_id(data,keuze)
    if viable == 0:
        print(f"{keuze} not in list")
    else:
        new_activiteit = input(f"Wat is de nieuwe activiteit van ID {keuze}? ")
        new_prijs = int(input(f"Wat is de prijs van {new_activiteit}? "))
        for row in data:
            if row[0] == keuze:
                row[2] = new_activiteit
                row[4] = new_prijs
        print(f"Activiteit van {keuze} is veranderd in {new_activiteit} en het kost {new_prijs}")
        return data

def sorteer_kostprijs(data,headers):
    t = sorted(data,key= lambda x:x[4], reverse=True)
    toon_data(t,headers)


def activiteiten_met_lunch(data,headers):
    filter_data = []
    for f in data:
        if f[5] == "Ja":
            filter_data.append(f)
    wegschrijven_excel(headers,filter_data)
    print("Excel aangemaakt")


def wegschrijven_excel(headers,source_data):
    # Create a new workbook for the combined data
    wb_combined = Workbook()
    combined_sheet = wb_combined.active
    filter_bestand = "activiteiten_lunch.xlsx"
    combined_sheet.append(headers)
    # Function to copy rows from a source sheet to the combined sheet with ID and store name
    for row in source_data:
        combined_sheet.append(list(row))
    wb_combined.save(filter_bestand)
    
def check_id(data,keuze):
    idlist = []
    for i, value in enumerate(data):
        idlist.append(data[i][0])
    if keuze in idlist:
        return 1
    else:
        return 0


data,headers = lees_excel(bestand)

functions_dict = {"1": toon_data, "2": voeg_data_toe, "3": verwijder_rij, "4": pas_locatie_aan, "5": pas_activiteit_aan, "6": sorteer_kostprijs, "7": activiteiten_met_lunch}

while True:
    print("1: Toon data"+"\n"
          "2: Voeg data toe"+"\n"
          "3: Verwijder rij"+"\n"
          "4: Pas locatie aan van activiteit"+"\n"
          "5: Pas activiteit aan van ID"+"\n"
          "6: Sorteer op kostprijs van duur naar goedkoop"+"\n"
          "7: Filter activiteiten met lunch"+"\n"
          "STOP: Stop controller")
    keuze = input("Wat wil je doen?").upper()
    if keuze == "STOP":
        break
    elif keuze in functions_dict.keys():
        functions_dict[keuze](data,headers)
    else:
        print("Foute ingave!")





#
# #Toon alle muzikanten van genre(x) in tabulate
# def toon_muzikanten_genre():
#     gefilterde_data = []
#     keuze = input(f"Geef het genre waarop je wilt filteren: ")
#     for row in data:
#         if keuze in row:
#             gefilterde_data.append(row)
#     print(tabulate(gefilterde_data, headers = headers))
#
# #Pas het instrument aan van de muzikant
# def pas_instrument_aan():
#     id = input("Geef het id waarvan je het instrument wilt aanpassen: ")
#     for row in data:
#         if id in row:
#             instrument = input(f"Wat is het instrument van dokter met id {id}? ")
#             row[2] = instrument
#             print(tabulate(data, headers= headers))
#
#
# #Sorteer alle muzikanten op naam a-z
# def sort_naam():
#     sorted_data = sorted(data, key = lambda x : x[1])
#     print(tabulate(sorted_data, headers=headers))
#
# #Schrijf huidige data weg naar een nieuw csv bestand
# def wegschrijven():
#     with open("muzikanten_data_update.csv", mode = 'w', newline='', encoding='utf-8') as file:
#         writer = csv.writer(file)
#         writer.writerow(headers)
#         writer.writerows(data)
#         print("\nDe bijgewerkte gegevens zijn succesvol opgeslagen in het bestand.")