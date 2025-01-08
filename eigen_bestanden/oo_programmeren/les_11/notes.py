import openpyxl

def load_workbook(file_path):
    """
    Load the workbook from the specified file path.
    """
    return openpyxl.load_workbook(file_path)

def toon_alle_kolom_namen(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolommen = [cell.value for cell in sheet[1]]
    print("Kolomnamen:", kolommen)
    workbook.close()
    return kolommen

def verwijder_kolom_met_naam(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolommen = toon_alle_kolom_namen(workbook_path)
    kolom_naam = input("Geef de naam van de kolom die verwijderd moet worden: ")

    if kolom_naam in kolommen:
        kolom_index = kolommen.index(kolom_naam) + 1
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            sheet.cell(row=row, column=kolom_index).value = None
        sheet.delete_cols(kolom_index)
        workbook.save(workbook_path)
        print(f"Kolom '{kolom_naam}' en de bijbehorende gegevens zijn verwijderd.")
    else:
        print(f"Kolom '{kolom_naam}' niet gevonden.")
    workbook.close()

def voeg_kolom_toe(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook.active
    kolom_naam = input("Geef de naam van de kolom die toegevoegd moet worden: ")

    sheet.insert_cols(sheet.max_column + 1)
    sheet.cell(row=1, column=sheet.max_column, value=kolom_naam)
    workbook.save(workbook_path)
    print(f"Kolom '{kolom_naam}' is toegevoegd.")
    workbook.close()

def toon_data_leeg(file_path):
    """
    Toon ID's van velden waar de cellen leeg zijn in de kolommen.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    leeg_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header
        id_val = row[0]
        if any(cell is None or str(cell).strip() == '' for cell in row):
            leeg_ids.append(id_val)

    print("ID's met lege velden:", leeg_ids)

# Bestandspad
bestand_pad = r'C:\Users\Jordan\Documents\SYNTRA\Python\SyntraPython\personeel1.xlsx'

# Voorbeeld gebruik
# while True:
#     print("\nKies een optie:")
#     print("1. Toon alle kolomnamen")
#     print("2. Verwijder een kolom met naam")
#     print("3. Voeg een kolom toe")
#     print("4. Stop")
#
#     keuze = input("Maak een keuze (1-4): ")
#
#     if keuze == '1':
#         toon_alle_kolom_namen(bestand_pad)
#     elif keuze == '2':
#         verwijder_kolom_met_naam(bestand_pad)
#     elif keuze == '3':
#         voeg_kolom_toe(bestand_pad)
#     elif keuze == '4':
#         print("Programma gestopt.")
#         break
#     else:
#         print("Ongeldige keuze, probeer opnieuw.")

toon_data_leeg(bestand_pad)
