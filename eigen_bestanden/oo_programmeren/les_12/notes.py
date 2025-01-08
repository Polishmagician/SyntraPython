from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load data from the original Excel files
file_hasselt = 'winkel_hasselt.xlsx'
file_genk = 'winkel_genk.xlsx'

wb_hasselt = load_workbook(file_hasselt)
wb_genk = load_workbook(file_genk)
sheet_hasselt = wb_hasselt.active
sheet_genk = wb_genk.active

# Create a new workbook for the combined data
wb_combined = Workbook()
combined_sheet = wb_combined.active
combined_sheet.title = "Winkels"

# Add headers with new columns "ID" and "Winkel"
headers = ["ID"] + [cell.value for cell in sheet_hasselt[1]] + ["Winkel"]
combined_sheet.append(headers)

# Function to copy rows from a source sheet to the combined sheet with ID and store name
def copy_rows_with_id(source_sheet, store_name, start_id):
    row_id = start_id
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        combined_sheet.append([row_id] + list(row) + [store_name])
        row_id += 1
    return row_id

# Copy data from both sheets
next_id = copy_rows_with_id(sheet_hasselt, "Hasselt", 1)
copy_rows_with_id(sheet_genk, "Genk", next_id)

# Save the combined workbook
combined_output_file = 'sales_data.xlsx'
wb_combined.save(combined_output_file)

# Print the output file path
print("Combined file saved at:", combined_output_file)
