from openpyxl import Workbook, load_workbook

bestand = 'teambuilding.xlsx'


#Open Excel bestand en lees in:
def lees_excel(bestand):
    wb_sheet_start =load_workbook(bestand)
    sheet_start = wb_sheet_start.active
    headers = [cell.value for cell in sheet_start[1]]
    data = []
    for row in sheet_start.iter_rows(min_row=2,values_only=True):
        data.append(list(row))
    return data,headers

def toon_data(data):
    for f in data:
        print(*f)

def voeg_data_toe(data):
    new_row = []
    new_id = data[-1][0]+1
    new_row.append(new_id)
    for f in headers[1:]:
        keuze = input(f"Wat is {f}? ")
        new_row.append(keuze)
    data.append(new_row)
    return data

def verwijder_rij(data):
    toon_data(data)
    keuze = int(input("Welke id wil je verwijderen?"))
    idlist = []
    for i,value in enumerate(data):
        idlist.append(data[i][0])
    if keuze in idlist:
        data.pop(keuze-1)
    else:
        print("not in list")
    return data


data,headers = lees_excel(bestand)
print(headers)

while True:
    print("1: Toon data"+"\n"
          "2: Stop"+"\n"
          "3: Voeg data toe"+"\n"
          "4: Verwijder rij")
    keuze = input("Wat wil je doen?")
    if keuze == "2":
        break
    elif keuze == "1":
        toon_data(data)
    elif keuze == "3":
        voeg_data_toe(data)
    elif keuze == "4":
        verwijder_rij(data)


# #Voeg een muzikant toe
# def nieuwe_muzikant():
#     nieuwe_id = str(int(data[-1][0])+1)
#     nieuw_muzikant = [nieuwe_id]
#     print("Voer onder deze lijn de gegevens in voor een nieuwe muzikant:")
#     for header in headers[1:]:
#         waarde = input(f"Voer {header} in: ")
#         nieuw_muzikant.append(waarde)
#     data.append(nieuw_muzikant)
#     print(tabulate(data, headers= headers))
#
# #Verwijder een muzikant
# def verwijder_muzikant():
#     id = input("Voer het id in dat je wilt verwijderen: ")
#     for index, row in enumerate(data):
#         if id in row:
#             data.pop(index)
#             print(tabulate(data, headers=headers))
#
#
#
# #Toon alle muzikanten van genre(x) in tabulate
# def toon_muzikanten_genre():
#     gefilterde_data = []
#     keuze = input(f"Geef het genre waarop je wilt filteren: ")
#     for row in data:
#         if keuze in row:
#             gefilterde_data.append(row)
#     print(tabulate(gefilterde_data, headers = headers))
#
# #Pas het instrument aan van de muzikant
# def pas_instrument_aan():
#     id = input("Geef het id waarvan je het instrument wilt aanpassen: ")
#     for row in data:
#         if id in row:
#             instrument = input(f"Wat is het instrument van dokter met id {id}? ")
#             row[2] = instrument
#             print(tabulate(data, headers= headers))
#
#
# #Sorteer alle muzikanten op naam a-z
# def sort_naam():
#     sorted_data = sorted(data, key = lambda x : x[1])
#     print(tabulate(sorted_data, headers=headers))
#
# #Schrijf huidige data weg naar een nieuw csv bestand
# def wegschrijven():
#     with open("muzikanten_data_update.csv", mode = 'w', newline='', encoding='utf-8') as file:
#         writer = csv.writer(file)
#         writer.writerow(headers)
#         writer.writerows(data)
#         print("\nDe bijgewerkte gegevens zijn succesvol opgeslagen in het bestand.")